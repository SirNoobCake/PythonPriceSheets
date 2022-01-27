from openpyxl import load_workbook

import os

def isfloat(num):
    try:
        float(num)
        return True
    except ValueError:
        return False

# Setting Year & Directory
directory = ""
year = 0
with open("year.txt", "r") as file_year:
    for line in file_year:
        for i in line.split(","):
            if isfloat(i.strip()):
                year = i
            else:
                directory = i
        
file_year.close()

#   Makes sure a file exists
filepath = os.path.join(directory, 'filename')
if not os.path.exists(directory):
    os.makedirs(directory)

#   Setting up vars for WSDP
rows_WSDP = 21
columns_WSDP = 19

products_WSDP_arr = [[0 for i in range(rows_WSDP)] for j in range(columns_WSDP)]

j = 0
k = -1 # Temp fix to make k == 0 for first iteration

#   Populating products_arr
with open("WSDP.txt", "r") as WSPD:
    for line in WSPD:
        for i in line.split(','):
            if isfloat(i.strip()):
                products_WSDP_arr[k][j] = float(i)
                if k == 1 & j == 15 | k > 12 & j == 15:
                    continue
                j += 1
            else:
                k += 1
                j = 0
WSPD.close()

#   Setting up vars for WSP
rows_WSP = 20
columns_WSP = 34

products_WSP_list = [[0 for i in range(rows_WSP)] for j in range(columns_WSP)]

#   Populating products_WSP_list
j = 0
k = -1

with open("WSP.txt", "r") as WSP:
    for line in WSP:
        for i in line.split(','):
            if isfloat(i.strip()):
                products_WSP_list[k][j] = float(i)
                if k == 1 & j == 15 | k > 12 & j == 15:
                    continue
                j += 1
            else:
                k += 1
                j = 0
WSP.close()

#   Setting up vars for RTPU
rows_RTPU = 6
columns_RTPU = 17

products_RTPU_list = [[0 for i in range(rows_RTPU)] for j in range(columns_RTPU)]

#   Populating products_RTPU_list
j = 0
k = -1

with open("RTPU.txt", "r") as RTPU:
    for line in RTPU:
        for i in line.split(','):
            if isfloat(i.strip()):
                products_RTPU_list[k][j] = float(i)
                if k == 1 & j == 15 | k > 12 & j == 15:
                    continue
                j += 1
            else:
                k += 1
                j = 0
RTPU.close()

# Setting up vars for RTD
rows_RTD = 20
columns_RTD = 19

products_RTD_arr = [[0 for i in range(rows_RTD)] for j in range(columns_RTD)]

j = 0
k = -1 # Temp fix to make k == 0 for first iteration

# Populating products_arr
with open("RTD.txt", "r") as RTD:
    for line in RTD:
        for i in line.split(','):
            if isfloat(i.strip()):
                products_RTD_arr[k][j] = float(i)
                if k == 1 & j == 15 | k > 12 & j == 15:
                    continue
                j += 1
            else:
                k += 1
                j = 0
RTD.close()

# k = product, 
#    0 = Delivery Price Mulch Poducts
#    1 = Delivery Price Soil Products
#    2 = Premium Bark
#    3 = Bark Blend
#    4 = Nature's Blend
#    5 = Beauty Bark
#    6 = Dyed Muclhes
#    7 = Safe Cover
#    8 = Clean Wood Chips
#    9 = Wood Chips
#    10 = Compost
#    11 = Leaf Compost
#    12 = Mushroom Soil
#  ****** Start of Changes in Yardages ****** 17
#    13 = Rain Garden Mix
#    14 = Screened Blend
#    15 = Screened Topsoil
#    16 = Regular Topsoil
#    17 = Fill Dirt
#    18 = Topsoil Overs

#   Setting up for Retail Delivered Prices Sheet

#   grab the active worksheet
load_RTD_wb = load_workbook("Retail Delivered Template.xlsx")
ws_RTD = load_RTD_wb.active

#   Set year in Excel
ws_RTD['R2'] = year + " Retail Delivered Prices"

#   Setting Prices
x = 6
for i in range(columns_RTD):
    j = 0
    if i == 1:
        for row in ws_RTD.iter_rows(min_col=5,min_row=29,max_col=24,max_row=29):
            for cell in row:
                cell.value = format(products_RTD_arr[i][j],'.2f')
                j += 1
                if j > 16:
                    break
        x = 6
    else:
        for row in ws_RTD.iter_rows(min_col=5,min_row=x,max_col=24,max_row=x):
            for cell in row:
                cell.value = format(products_RTD_arr[i][j],'.2f')
                j += 1
                if i > 12:
                    if j > 16:
                        break

    # This Makes the Program only interact with the desired collums
    if x == 6:
        x += 1
    else:
        if x == 27:
            x += 3
        else:
            x += 2

# Save Worksheet
load_RTD_wb.save(directory + "\\" + year + "_Retail_Delivered_Prices.xlsx")



#   Setting up for Retail Picked-up Worksheet
load_RTPU_wb = load_workbook("Retail Picked-up Template.xlsx")
ws_RTPU = load_RTPU_wb.active

#   Set Year in Sheet
ws_RTPU['H1'] = year + " Retail"

#   Setting Prices
x = 0
for i in range(columns_RTPU):
    k = 0
    j = 0
    for row in ws_RTPU.iter_rows(min_col=5,min_row=x+6,max_col=11,max_row=x+6):
        for cell in row:
            k += 1
            if k == 6:
                continue
            cell.value = format(products_RTPU_list[i][j], '.2f')
            j += 1
            if j == 5:
               if i > 10 and i < 15:
                   break
        x += 2

load_RTPU_wb.save(directory + "\\" + year + "_Retail_Picked-up_Price.xlsx")



#   Setting up for Wholesale Prices Sheet
load_WSP_wb = load_workbook("Wholesale PriceSheet Template.xlsx")
ws_WSP = load_WSP_wb.active

#   Set Year in sheet
ws_WSP['S1'] = year + " Wholesale Prices"

#   Setting Prices
x = 0
for i in range(columns_WSP):
    j = 0
    if i < 17:
        for row in ws_WSP.iter_rows(min_col=4,min_row=x + 5,max_col=7,max_row=x + 5):
            for cell in row:
                cell.value = format(products_WSP_list[i][j],'.2f')
                j += 1
                if j > 4:
                    break
    else:
        if i == 17:
            x = 0
        for row in ws_WSP.iter_rows(min_col=9,min_row=x + 5,max_col=28,max_row=x + 5):
            for cell in row:
                cell.value = format(products_WSP_list[i][j],'.2f')
                j += 1
                if i > 27:
                    if j > 15:
                        break
    if x == 3 or x == 5 or x == 12:
        x += 2
    else:
        x += 1
load_WSP_wb.save(directory + "\\" + year + "_Wholesale_Prices.xlsx")

#   Setting up for Wholesale Delivered Prices Sheet

#   grab the active worksheet
load_WSDP_wb = load_workbook("Wholesale Delivered Template.xlsx")
ws_WSDP = load_WSDP_wb.active

#   Set year in Excel
ws_WSDP['R2'] = year + " Wholesale Delivered Prices"

#   Setting Prices
x = 6
for i in range(columns_WSDP):
    j = 0
    if i == 1:
        for row in ws_WSDP.iter_rows(min_col=5,min_row=29,max_col=25,max_row=29):
            for cell in row:
                cell.value = format(products_WSDP_arr[i][j],'.2f')
                j += 1
                if j > 16:
                    break
        x = 6
    else:
        for row in ws_WSDP.iter_rows(min_col=5,min_row=x,max_col=25,max_row=x):
            for cell in row:
                cell.value = format(products_WSDP_arr[i][j],'.2f')
                j += 1
                if i > 12:
                    if j > 16:
                        break

    # This Makes the Program only interact with the desired collums
    if x == 6:
        x += 1
    else:
        if x == 27:
            x += 3
        else:
            x += 2

# Save Worksheet
load_WSDP_wb.save(directory + "\\" + year + "_Wholesale_Delivered_Prices.xlsx")

